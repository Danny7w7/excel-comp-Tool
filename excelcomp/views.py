from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render, HttpResponse

from .models import *

from io import BytesIO

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import os
import math

from datetime import datetime, timedelta


# Create your views here.

# Auth
def login_(request):
    if request.user.is_authenticated:
        return redirect(index)
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect(index)
        else:
            msg = 'Wrong data, try again'
            return render(request, 'login.html', {'msg':msg})
    else:
        return render(request, 'login.html')
    
def logout_(request):
    logout(request)
    return redirect(index)


@login_required(login_url='/login')
def index(request):
    if request.method == 'POST':
        dfs = []
        if request.POST['report'] == 'client_report':
            files = [request.FILES['file1'], request.FILES['file2']]
        elif request.POST['report'] == 'broker_report' or request.POST['report'] == 'oneill_report':
            files = [request.FILES['file1']]
        elif request.POST['report'] == 'comparative_report':
            return comparative_report()
        for file in files:
            extension = get_extension(file)
            if extension == '.xlsx':
                excel_file = pd.ExcelFile(file)
                dfs.append(pd.read_excel(file))
            else:
                dfs.append(pd.read_csv(file, delimiter=','))
        return generate_excel(dfs, request, get_last_8_letters(request.FILES['file1'].name), file)
    return render(request, 'index.html')

def show_comparative(request):
    supps = InsxCloudSupp.objects.all()
    exceptions = []
    policyNumberDict = {}

    for supp in supps:
        if supp.issuerPolicyNumber == '':
            exceptions.append(supp)
        else:
            if supp.issuerPolicyNumber not in policyNumberDict:
                policyNumberDict[supp.issuerPolicyNumber] = []
            policyNumberDict[supp.issuerPolicyNumber].append(supp)

    current_monday, previous_monday = get_two_mondays(datetime.now())
    clients = classify_policies(policyNumberDict, current_monday, previous_monday)
    context = {
        'clients': clients,
        'current_monday': current_monday,
        'previous_monday': previous_monday
    }
    return render(request, 'show_comparative.html', context)

def client_report(df):
    df.columns = df.columns.str.strip()
    filtered_df = df[df['Provider Status'] != 'Active']
    remaining_df = df[df['Provider Status'] == 'Active']
    for index, row in df.iterrows():
        supp = InsxCloudSupp()

        if pd.isna(row['Issuer Policy Number']) or math.isnan(row['Issuer Policy Number']):
            supp.issuerPolicyNumber = None
        else:
            supp.issuerPolicyNumber = int(row['Issuer Policy Number'])

        print(supp.issuerPolicyNumber)
        supp.status = row['Provider Status']
        supp.firstName = row['First Name']
        supp.lastName = row['Last Name']
        supp.middleName = row['Middle Name']
        supp.uploadDate = datetime.now()
        supp.dateSubmit = datetime.strptime(row['Date Submitted'], "%m/%d/%Y")
        supp.dateEffective = datetime.strptime(row['Effective Date'], "%m/%d/%Y")
        supp.dateCancellation = datetime.strptime(row['Cancellation Date'], "%m/%d/%Y")
        supp.gender = row['Gender']
        supp.address = row['Address']
        supp.city = row['City']
        supp.state = row['State']
        supp.zipCode = row['Zip Code']
        supp.country = row['County']
        supp.phoneNumber = row['Phone Number']
        supp.email = row['Email']
        supp.lineOfCoverage = row['Line Of Coverage']
        supp.insuranceCompany = row['Insurance Company']
        supp.agencyName = row['Agency Name']
        supp.broker = row['Producer']
        supp.npn = row['NPN']
        supp.transactionId = row['Transaction Id']
        supp.save()
    return remaining_df, filtered_df

def broker_report(df):
    date_columns = ['Received Date', 'Issue Date', 'Effective Date', 'Paid to Date', 'Subscriber DOB']

    # Estandarizar las fechas en las columnas especificadas al formato MM/DD/YYYY
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%m/%d/%Y')


    df.columns = df.columns.str.strip()
    active = df[df['Status'] == 'Active']
    active_grace_period = df[df['Status'] == 'Active (Grace Period) (Payment Error)']
    void_return = df[df['Status'] == 'Void-Return-Existing Insurance']
    active_but_date = df[(df['Status'] == 'Active (Payment Error)') & (df['Paid to Date'].notna()) & (df['Paid to Date'] != '')]
    active_but_no_date = df[(df['Status'] == 'Active (Payment Error)') & (df['Paid to Date'].isna() | (df['Paid to Date'] == ''))]
    automatic_termination = df[df['Status'] == 'Automatic Termination (Payment Error)']
    other_case = df[(df['Status'].notna()) &
                    (df['Status'] != '') &
                    (df['Status'] != 'Active') & 
                    (df['Status'] != 'Active (Payment Error)') & 
                    (df['Status'] != 'Active (Grace Period) (Payment Error)') & 
                    (df['Status'] != 'Void-Return-Existing Insurance') &
                    (df['Status'] != 'Automatic Termination (Payment Error)')]


    return active, active_grace_period, void_return, active_but_date, active_but_no_date, automatic_termination, other_case

def oneill_report(file):

    base_file = 'oneill_report_base.xlsx'
    excel_file = pd.ExcelFile(file)

    # Crear un nuevo workbook para escribir los resultados
    wb = load_workbook(base_file)
    ws = wb.active

    # Fila de inicio para copiar datos
    start_row = 2
    cont = 0
    
    for i in range(len(excel_file.sheet_names)):

        print(f'Voy por la pagina numero {i} - {excel_file.sheet_names[i]}')
        sheet = excel_file.sheet_names[i]

        # Cargar la primera hoja y quitar espacios en los nombres de las columnas
        df_sheet = pd.read_excel(file, sheet_name=sheet)
        df_sheet.columns = df_sheet.columns.str.strip()

        # Función para verificar si un PolicyNumber ya está en la columna 5
        def policy_exists(ws, policy_number):
            for cell in ws['E']:  # Columna 5 es la columna 'E'
                if cell.row >= 2 and cell.value == policy_number:
                    return True
            return False

        # Iterar sobre los PolicyNumber de la primera hoja
        for policyNumber in df_sheet['PolicyNumber']:
            cont += 1
            print(f'ITERACION NUMERO {cont}')
            if pd.isna(policyNumber):
                continue

            # Verificar si el PolicyNumber ya está en la columna 5 del Excel
            if policy_exists(ws, policyNumber):
                print(f'El PolicyNumber {policyNumber} ya ha sido agregado anteriormente. ITERACION: {cont}')
                continue
            
            # Filtrar solo las columnas A a G (1 a 7)
            row_data = df_sheet[df_sheet['PolicyNumber'] == policyNumber].iloc[0].values[:7]
            
            # Copiar las columnas seleccionadas al archivo base
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=start_row, column=col_num, value=cell_value)
            start_row += 1

            # Comprobar si el PolicyNumber está en las demás hojas
            print(f'Comprobando {policyNumber}. ITERACION: {cont}')
            for sheet_name in excel_file.sheet_names[i+1:]:
                df_other_sheet = pd.read_excel(file, sheet_name=sheet_name)
                df_other_sheet.columns = df_other_sheet.columns.str.strip()
                
                if policyNumber in df_other_sheet['PolicyNumber'].values:
                    print(f'{policyNumber} SI está en {sheet_name}')
                    
                    # Obtener el `StatementDate` de la fila correspondiente
                    statement_date = df_other_sheet[df_other_sheet['PolicyNumber'] == policyNumber].iloc[0]['StatementDate']

                    # Formatear la fecha como MM/DD/YYYY
                    statement_date = pd.to_datetime(statement_date).strftime('%m/%d/%Y')

                    # Obtener el `CoverageMonth`
                    coverage_month = df_other_sheet[df_other_sheet['PolicyNumber'] == policyNumber].iloc[0]['CoverageMonth']
                    try:
                        month_number = pd.to_datetime(coverage_month).month
                    except:
                        month_number = 0
                        print(f'En la juega que esta vaina dio error, CoverageMonth: {coverage_month}')

                    # Colocar el `StatementDate` en la columna correspondiente al número del mes más 8
                    target_column = month_number + 8
                    ws.cell(row=start_row-1, column=target_column, value=statement_date)

    # Guardar el archivo modificado
    wb.save('oneill_report_output.xlsx')
    print(f'Reporte generado en: oneill_report_output.xlsx')

def comparative_report():
    exceptions = []
    report = []
    policyNumberDict = {}
    supps = InsxCloudSupp.objects.all()

    for supp in supps:
        if supp.issuerPolicyNumber == '':
            exceptions.append(supp)
        else:
            if supp.issuerPolicyNumber not in policyNumberDict:
                policyNumberDict[supp.issuerPolicyNumber] = []
            policyNumberDict[supp.issuerPolicyNumber].append(supp)
    
    for policyNumber, supp_group in policyNumberDict.items():
        report_item = {
            'PolicyNumber': policyNumber,
            'broker': supp_group[0].broker,
            'company': supp_group[0].agencyName,
            'firstNames': supp_group[0].firstName,
            'lastName': supp_group[0].lastName,
        }
        # Agregar los estados para cada lunes
        monday_dates = get_mondays(datetime(2024, 7, 15), datetime(2024, 11, 1))
        for date in monday_dates:
            report_item[date] = ''  # Valor por defecto
        
        for supp in supp_group:
            supp_date = supp.uploadDate.strftime("%Y-%m-%d")
            if supp_date in monday_dates:
                report_item[supp_date] = supp.status
                
        report.append(report_item)

    # Crear un DataFrame a partir de la lista de diccionarios
    df = pd.DataFrame(report)

    # Crear un libro de trabajo de Excel con openpyxl
    workbook = Workbook()
    sheet = workbook.active

    # Escribir los encabezados
    for idx, col_name in enumerate(df.columns, 1):
        sheet.cell(row=1, column=idx, value=col_name)

    # Escribir los datos
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Definir los estilos
    green_fill = PatternFill(start_color="b6d7a8", end_color="b6d7a8", fill_type="solid")
    red_fill = PatternFill(start_color="ea9999", end_color="ea9999", fill_type="solid")

    # Aplicar los estilos condicionales
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=sheet.max_column):  # Empieza en la columna 6, donde están los lunes
        for cell in row:
            if cell.value == 'Active':
                cell.fill = green_fill
            elif cell.value == 'Cancelled' or cell.value == 'Terminated':
                cell.fill = red_fill

    # Guardar el archivo en un BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Crear la respuesta HTTP
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Comparative InsxCloud Report.xlsx"'

    return response

def generate_excel(dfs, request, nameUwU, file=None):
    today = datetime.now()
    format_date = today.strftime("%m%d%y")

    remaining_dfs = []
    filtered_dfs = []
    actives = []
    actives_grace_period = []
    voids_return = []
    actives_but_date = []
    actives_but_no_date = []
    automatic_terminations = []
    others_case = []

    if request.POST['report'] == 'client_report':
        for df in dfs:
            remaining_df, filtered_df = client_report(df)
            remaining_dfs.append(remaining_df)
            filtered_dfs.append(filtered_df)
    elif request.POST['report'] == 'broker_report':
        for df in dfs:
            active, active_grace_period, void_return, active_but_date, active_but_no_date, automatic_termination, other_case = broker_report(df)
            actives.append(active)
            actives_grace_period.append(active_grace_period)
            voids_return.append(void_return)
            actives_but_date.append(active_but_date)
            actives_but_no_date.append(active_but_no_date)
            automatic_terminations.append(automatic_termination)
            others_case.append(other_case)
    elif request.POST['report'] == 'oneill_report':
        for df in dfs:
            oneill_report(file)
        return redirect(index)


    # Crear un nuevo archivo Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Escribir los DataFrames originales y filtrados en nuevas hojas
        
        if request.POST['report'] == 'client_report':
            sheet_names = {
                0: ("ACTIVE LAP & TRU", "CANCELADOS LAP & TRU"),
                1: ("ACTIVE SECURE", "CANCELADOS SECURE")
            }
            for i, (remaining_df, filtered_df) in enumerate(zip(remaining_dfs, filtered_dfs)):
                if not remaining_df.empty:
                    remaining_df.to_excel(writer, sheet_name=sheet_names[i][0], index=False)
                if not filtered_df.empty:
                    filtered_df.to_excel(writer, sheet_name=sheet_names[i][1], index=False)
                    
        elif request.POST['report'] == 'broker_report':
            sheet_names = {
                0: "Active Clients",
                1: "Actives Clients with payment",
                2: "Actives Clients Without payment",
                3: "Actives Clients in grace period",
                4: "Automatic Termination",
                5: "Void return",
                6: "Other cases",
            }
            for i, (active_client, active_grace_period, void_return, active_but_date, active_but_no_date, automatic_termination, other_case) in enumerate(zip(actives, actives_grace_period, voids_return, actives_but_date, actives_but_no_date, automatic_terminations, others_case)):
                if not active_client.empty:
                    active_client.to_excel(writer, sheet_name=sheet_names[0], index=False)
                if not active_grace_period.empty:
                    active_grace_period.to_excel(writer, sheet_name=sheet_names[3], index=False)
                if not void_return.empty:
                    void_return.to_excel(writer, sheet_name=sheet_names[5], index=False)
                if not active_but_date.empty:
                    active_but_date.to_excel(writer, sheet_name=sheet_names[1], index=False)
                if not active_but_no_date.empty:
                    active_but_no_date.to_excel(writer, sheet_name=sheet_names[2], index=False)
                if not automatic_termination.empty:
                    automatic_termination.to_excel(writer, sheet_name=sheet_names[4], index=False)
                if not other_case.empty:
                    other_case.to_excel(writer, sheet_name=sheet_names[6], index=False)
        
        
        # Verificar si hay hojas en el archivo
        if len(writer.sheets) == 0:
            raise ValueError("No se pudo crear el archivo Excel, ya que no hay hojas visibles.")

    output.seek(0)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if request.POST['report'] == 'client_report':
        response['Content-Disposition'] = f'attachment; filename="Insxcloud - Client Report {format_date}.xlsx"'
    if request.POST['report'] == 'broker_report':
        response['Content-Disposition'] = f'attachment; filename="ExcelComp - Broker Report {format_date}.xlsx"'

    return response

def is_cancelled_or_terminated(status):
    return status in ["Cancelled", "Terminated", None]

def classify_policy(s1_status, s2_status):
    if s1_status is None and s2_status == "Active":
        return "Active_without_policynumber"
    elif is_cancelled_or_terminated(s1_status) and s2_status == "Active":
        return "New_Canceled"
    elif s1_status == "Active" and is_cancelled_or_terminated(s2_status):
        return "Recovered"
    elif is_cancelled_or_terminated(s1_status) and is_cancelled_or_terminated(s2_status):
        return "Cancelled_old"
    elif s1_status == "Active" and s2_status == "Active":
        return "Active"
    else:
        return "Unclassified"

def classify_policies(policyNumberDict, current_monday, previous_monday):
    classification = {
        "New_Canceled": [],
        "Recovered": [],
        "Cancelled_old": [],
        "Active": [],
        "Active_without_policynumber": [],
        "Unclassified": []
    }
    for policy_number, supps in policyNumberDict.items():
        current_supp = next((supp for supp in supps if supp.uploadDate == current_monday), None)
        previous_supp = next((supp for supp in supps if supp.uploadDate == previous_monday), None)
        
        s1_status = current_supp.status if current_supp else None
        s2_status = previous_supp.status if previous_supp else None
        
        category = classify_policy(s1_status, s2_status)
        
        client_info = {
            'policy_number': policy_number,
            'broker': current_supp.broker if current_supp else (previous_supp.broker if previous_supp else ''),
            'agency_name': current_supp.agencyName if current_supp else (previous_supp.agencyName if previous_supp else ''),
            'first_name': current_supp.firstName if current_supp else (previous_supp.firstName if previous_supp else ''),
            'last_name': current_supp.lastName if current_supp else (previous_supp.lastName if previous_supp else ''),
            'previous_status': s2_status or 'Not policy',
            'current_status': s1_status or 'Not policy'
        }
        
        classification[category].append(client_info)
        
    return classification

def get_extension(file):
    file_name, extension = os.path.splitext(file.name)
    return extension

def get_last_8_letters(text):
  inicio = len(text) - 12
  final = len(text) -4
  last_8_letters = text[inicio:final]
  return last_8_letters

def get_mondays(start_date, end_date):
    current_date = start_date
    mondays = []
    
    while current_date <= end_date:
        if current_date.weekday() == 0:  # 0 representa el lunes
            mondays.append(current_date.strftime("%Y-%m-%d"))
        current_date += timedelta(days=1)
    
    return mondays

def get_two_mondays(date):
    day_of_week = date.weekday()
    days_to_monday = timedelta(days=-day_of_week)
    current_monday = date + days_to_monday
    previous_monday = current_monday - timedelta(days=7)
    
    return current_monday.date(), previous_monday.date()